import os

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import copy
from typing import NamedTuple


def load_ub_value(folder_path, instance_name):
    """Charge la valeur UB depuis le fichier instance.UB"""
    ub_file = os.path.join(folder_path, f"{instance_name}.UB")
    try:
        with open(ub_file, 'r') as f:
            return float(f.read().strip())
    except (FileNotFoundError, ValueError):
        return None

def format_pct(val):
    if val is None:
        return "-"
    elif abs(val) < 1:
        return val
    else:
        return int(round(val))


# --- CSV filenames / signatures used across the workbook and LaTeX aggregation ---

SOL_ALL_14400 = "Sol_All_14400.csv"
SOL_ALL_3600 = "Sol_All_3600.csv"
SOL_CPLEX_14400 = "Sol_CPLEX_14400.csv"
SOL_CPLEX_3600 = "Sol_CPLEX_3600.csv"
FF_SUFFIX_14400 = "-FF=1_-TL=14400.csv"
CPL_MARKER_LONG_RUN = "-CPL=1_-TL=14400.csv"
CPL_MARKER_IN_SHORT_NAME = "-CPL=1.csv"

# Long heuristic run name fragment (discovery pass uses substring match)
SHORT_HEURISTIC_RUN_TAG = (
    "Sol-FR=1_-BC=2_-IC=2_-GAP=2_-MS=20_-SC=0_-ACO=0_-H=2_-YT=2_-"
    "AI=0_-SFC=0_-FF=1_-TL=3601"
)
SHORT_HEURISTIC_RUN_FILE = SHORT_HEURISTIC_RUN_TAG + ".csv"

SOL_IC_3600 = "Sol_All-IC_3600.csv"
SOL_MS_3600 = "Sol_ALL-MS_3600.csv"
SOL_RS_3600 = "Sol_All-RS_3600.csv"
SOL_GAP_VARIANT_3600 = "Sol_All-GAP_3600.csv"
SOL_FF_VARIANT_3600 = "Sol_All-FF_3600.csv"
SOL_DEFAULT_BC1 = "Sol_Default_3600.csv"
SOL_LEGACY_DEFAULT_BC1 = (
    "Sol-FR=1_-BC=1_-IC=0_-GAP=0_-MS=0_-SC=0_-ACO=0_-H=0_-YT=0_-"
    "AI=0_-SFC=0_-FF=0_-TL=3601.csv"
)


# Raw instance folder -> (human-readable LaTeX label, bucket 0..2 for "10 / 20 / 30 customers")
INSTANCE_FOLDER_META = {
    "Instance_2P_10C_2H_5PE": ("(2,10,2,5)", 0),
    "Instance_2P_20C_2H_5PE": ("(2,20,2,5)", 1),
    "Instance_2P_30C_2H_5PE": ("(2,30,2,5)", 2),
    "Instance_4P_10C_2H_5PE": ("(4,10,2,5)", 0),
    "Instance_4P_20C_2H_5PE": ("(4,20,2,5)", 1),
    "Instance_4P_30C_2H_5PE": ("(4,30,2,5)", 2),
    "Instance_2P_10C_1H_5PE": ("(2,10,1,5)", 0),
    "Instance_2P_20C_1H_5PE": ("(2,20,1,5)", 1),
    "Instance_2P_30C_1H_5PE": ("(2,30,1,5)", 2),
    "Instance_4P_10C_1H_5PE": ("(4,10,1,5)", 0),
    "Instance_4P_20C_1H_5PE": ("(4,20,1,5)", 1),
    "Instance_4P_30C_1H_5PE": ("(4,30,1,5)", 2),
    "Instance_6P_10C_1H_5PE": ("(6,10,1,5)", 0),
    "Instance_6P_20C_1H_5PE": ("(6,20,1,5)", 1),
    "Instance_6P_30C_1H_5PE": ("(6,30,1,5)", 2),
    "Instance_6P_10C_2H_5PE": ("(6,10,2,5)", 0),
    "Instance_6P_20C_2H_5PE": ("(6,20,2,5)", 1),
    "Instance_6P_30C_2H_5PE": ("(6,30,2,5)", 2),
    "Instance_2P_10C_1H_10PE": ("(2,10,1,10)", 0),
    "Instance_2P_20C_1H_10PE": ("(2,20,1,10)", 1),
    "Instance_2P_30C_1H_10PE": ("(2,30,1,10)", 2),
    "Instance_4P_10C_1H_10PE": ("(4,10,1,10)", 0),
    "Instance_4P_20C_1H_10PE": ("(4,20,1,10)", 1),
    "Instance_4P_30C_1H_10PE": ("(4,30,1,10)", 2),
    "Instance_6P_10C_1H_10PE": ("(6,10,1,10)", 0),
    "Instance_6P_20C_1H_10PE": ("(6,20,1,10)", 1),
    "Instance_6P_30C_1H_10PE": ("(6,30,1,10)", 2),
    "Instance_2P_10C_2H_10PE": ("(2,10,2,10)", 0),
    "Instance_2P_20C_2H_10PE": ("(2,20,2,10)", 1),
    "Instance_2P_30C_2H_10PE": ("(2,30,2,10)", 2),
    "Instance_4P_10C_2H_10PE": ("(4,10,2,10)", 0),
    "Instance_4P_20C_2H_10PE": ("(4,20,2,10)", 1),
    "Instance_4P_30C_2H_10PE": ("(4,30,2,10)", 2),
    "Instance_6P_10C_2H_10PE": ("(6,10,2,10)", 0),
    "Instance_6P_20C_2H_10PE": ("(6,20,2,10)", 1),
    "Instance_6P_30C_2H_10PE": ("(6,30,2,10)", 2),
}


# Ordre fixe du parcours (cohérent avec les clés de INSTANCE_FOLDER_META)
INSTANCE_FOLDER_ORDER = [
    'Instance_2P_10C_1H_5PE', 'Instance_2P_20C_1H_5PE', 'Instance_2P_30C_1H_5PE',
    'Instance_4P_10C_1H_5PE', 'Instance_4P_20C_1H_5PE', 'Instance_4P_30C_1H_5PE',
    'Instance_6P_10C_1H_5PE', 'Instance_6P_20C_1H_5PE', 'Instance_6P_30C_1H_5PE',
    'Instance_2P_10C_2H_5PE', 'Instance_2P_20C_2H_5PE', 'Instance_2P_30C_2H_5PE',
    'Instance_4P_10C_2H_5PE', 'Instance_4P_20C_2H_5PE', 'Instance_4P_30C_2H_5PE',
    'Instance_6P_10C_2H_5PE', 'Instance_6P_20C_2H_5PE', 'Instance_6P_30C_2H_5PE',
    'Instance_2P_10C_1H_10PE', 'Instance_2P_20C_1H_10PE', 'Instance_2P_30C_1H_10PE',
    'Instance_4P_10C_1H_10PE', 'Instance_4P_20C_1H_10PE', 'Instance_4P_30C_1H_10PE',
    'Instance_6P_10C_1H_10PE', 'Instance_6P_20C_1H_10PE', 'Instance_6P_30C_1H_10PE',
    'Instance_2P_10C_2H_10PE', 'Instance_2P_20C_2H_10PE', 'Instance_2P_30C_2H_10PE',
    'Instance_4P_10C_2H_10PE', 'Instance_4P_20C_2H_10PE', 'Instance_4P_30C_2H_10PE',
    'Instance_6P_10C_2H_10PE', 'Instance_6P_20C_2H_10PE', 'Instance_6P_30C_2H_10PE',
]


class SolutionCsvFilenames(NamedTuple):
    long_ff_csv: str | None
    short_ff_csv: str | None
    long_cplex_csv: str | None
    short_cplex_csv: str | None


def _resolve_solution_csv_filenames(csv_files) -> SolutionCsvFilenames:
    """Repère dans un dossier les CSV All/FF (court / long TL) et CPLEX associés."""
    long_ff_csv = short_ff_csv = long_cplex_csv = short_cplex_csv = None
    for csv_file in csv_files:
        if _pick_ff_long_csv(csv_file):
            long_ff_csv = csv_file
        elif _pick_ff_short_csv(csv_file):
            short_ff_csv = csv_file
        if _pick_cplex_long_csv(csv_file):
            long_cplex_csv = csv_file
        elif _pick_cplex_short_csv(csv_file):
            short_cplex_csv = csv_file
    return SolutionCsvFilenames(long_ff_csv, short_ff_csv, long_cplex_csv, short_cplex_csv)


INSTANCES_PER_BUCKET = 12  # Dossiers par bande « 10 / 20 / 30 clients » dans les tables LaTeX


def _inject_ub_derived_columns(
    folder_path, df, is_long_ff_tl: bool, is_short_ff_tl: bool
):
    """
    Construit la colonne Upper et, si pertinent, UB / déviation % / gap UB–Lower ;
    mute ``df`` pour les lignes All/FF. Retourne (upper_column, +compteur UB long court).
    """
    ub_column = []
    deviation_column = []
    gap_ub_lower_column = []
    upper_column = []
    delta_long = delta_short = 0
    for _, row in df.iterrows():
        upper_value = float(row["Upper"])
        upper_column.append(upper_value)
        if not (is_long_ff_tl or is_short_ff_tl):
            continue

        instance_name = row.iloc[0]
        ub_value = load_ub_value(folder_path, instance_name)

        if pd.isna(ub_value):
            ub_column.append(None)
            deviation_column.append(None)
            gap_ub_lower_column.append(None)
            continue

        ub_value = float(ub_value)

        if upper_value < 10000 and ub_value < 10000 and upper_value != 0:
            deviation_pct = ((ub_value - upper_value) / upper_value) * 100
        else:
            deviation_pct = None

        ub_column.append(ub_value)
        deviation_column.append(deviation_pct)

        if ub_value is not None and not pd.isna(row["Lower"]) and ub_value < 10000:
            lower_value = float(row["Lower"])
            gap_ub_lower_pct = (
                ((ub_value - lower_value) / lower_value) * 100
                if lower_value != 0
                else None
            )
        else:
            gap_ub_lower_pct = None

        gap_ub_lower_column.append(gap_ub_lower_pct)

        if is_long_ff_tl and ub_value < 10000:
            delta_long += 1
        if is_short_ff_tl and ub_value < 10000:
            delta_short += 1

    if is_long_ff_tl or is_short_ff_tl:
        df["UB"] = ub_column
        df["Deviation %"] = deviation_column
        df["Gap UB-Lower %"] = gap_ub_lower_column

    return upper_column, delta_long, delta_short


def _style_instance_sheet(ws):
    """Accentue séparateurs, lignes Stat: et sections comparatives."""
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            if val.startswith("---"):
                cell.font = Font(bold=True)
            elif val.startswith("Stat:"):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
            elif "RÉCAPITULATIF" in val or "COMPARAISON" in val:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(
                    start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
                )


def _append_latex_table7_aggregate_rows(
    table: str,
    *,
    aver_opt_all,
    aver_gap_all,
    aver_ubgap_all,
    aver_opt_ic,
    aver_gap_ic,
    aver_ubgpa_ic,
    aver_opt_gap,
    aver_gap_gap,
    aver_ubgap_gap,
    aver_opt_ms,
    aver_gap_ms,
    aver_ubgap_ms,
    aver_opt_rs,
    aver_gap_rs,
    aver_ubgap_rs,
    aver_opt_ff,
    aver_gap_ff,
    aver_ubgap_ff,
    aver_opt_def,
    aver_gap_def,
):
    """Ajoute les lignes d'agrégation (méthodes × trois tailles)."""
    d = INSTANCES_PER_BUCKET
    table += (
        f"All & {aver_opt_all[0]} & {aver_gap_all[0]/d:.3f} & {aver_ubgap_all[0]/d:.3f} "
        f"& {aver_opt_all[1]} & {aver_gap_all[1]/d:.3f} & {aver_ubgap_all[1]/d:.3f} "
        f"& {aver_opt_all[2]} & {aver_gap_all[2]/d:.3f} & {aver_ubgap_all[2]/d:.3f} \\\\  \n"
        f"IC & {aver_opt_ic[0]} & {aver_gap_ic[0]/d:.3f} & {aver_ubgpa_ic[0]/d:.3f} "
        f"& {aver_opt_ic[1]} & {aver_gap_ic[1]/d:.3f} & {aver_ubgpa_ic[1]/d:.3f} "
        f"& {aver_opt_ic[2]} & {aver_gap_ic[2]/d:.3f} & {aver_ubgpa_ic[2]/d:.3f} \\\\ \n"
        f"GAP & {aver_opt_gap[0]} & {aver_gap_gap[0]/d:.3f} & {aver_ubgap_gap[0]/d:.3f} "
        f"& {aver_opt_gap[1]} & {aver_gap_gap[1]/d:.3f} & {aver_ubgap_gap[1]/d:.3f} "
        f"& {aver_opt_gap[2]} & {aver_gap_gap[2]/d:.3f} & {aver_ubgap_gap[2]/d:.3f} \\\\ \n"
        f"MS & {aver_opt_ms[0]} & {aver_gap_ms[0]/d:.3f} & {aver_ubgap_ms[0]/d:.3f} "
        f"& {aver_opt_ms[1]} & {aver_gap_ms[1]/d:.3f} & {aver_ubgap_ms[1]/d:.3f} "
        f"& {aver_opt_ms[2]} & {aver_gap_ms[2]/d:.3f} & {aver_ubgap_ms[2]/d:.3f} \\\\ \n"
        f"RS & {aver_opt_rs[0]} & {aver_gap_rs[0]/d:.3f} & {aver_ubgap_rs[0]/d:.3f} "
        f"& {aver_opt_rs[1]} & {aver_gap_rs[1]/d:.3f} & {aver_ubgap_rs[1]/d:.3f} "
        f"& {aver_opt_rs[2]} & {aver_gap_rs[2]/d:.3f} & {aver_ubgap_rs[2]/d:.3f} \\\\ \n"
        f"FF & {aver_opt_ff[0]} & {aver_gap_ff[0]/d:.3f} & {aver_ubgap_ff[0]/d:.3f} "
        f"& {aver_opt_ff[1]} & {aver_gap_ff[1]/d:.3f} & {aver_ubgap_ff[1]/d:.3f} "
        f"& {aver_opt_ff[2]} & {aver_gap_ff[2]/d:.3f} & {aver_ubgap_ff[2]/d:.3f} \\\\ \n"
        f"DEF & {aver_opt_def[0]} & {aver_gap_def[0]/d:.3f} & - & {aver_opt_def[1]} "
        f"& {aver_gap_def[1]/d:.3f} & - & {aver_opt_def[2]} & {aver_gap_def[2]/d:.3f} & - \\\\ \\hline \n"
        "\\end{tabular} \n\\end{table} \n"
    )
    return table


def _latex_close_outer_table(fragment: str) -> str:
    return fragment + "\\hline \n\\end{tabular} \n\\end{table} \n"


def _build_latex_table3_runtime_totals(
    *,
    aver_ite10_Short,
    aver_ite10_Long,
    aver_ite20_Short,
    aver_ite20_Long,
    aver_ite30_Short,
    aver_ite30_Long,
    aver_MSolving10_Short,
    aver_MSolving10_Long,
    aver_MSolving20_Short,
    aver_MSolving20_Long,
    aver_MSolving30_Short,
    aver_MSolving30_Long,
    aver_Subsolving10_Short,
    aver_Subsolving10_Long,
    aver_Subsolving20_Short,
    aver_Subsolving20_Long,
    aver_Subsolving30_Short,
    aver_Subsolving30_Long,
    aver_BendersCuts10_Short,
    aver_BendersCuts10_Long,
    aver_BendersCuts20_Short,
    aver_BendersCuts20_Long,
    aver_BendersCuts30_Short,
    aver_BendersCuts30_Long,
) -> str:
    d = INSTANCES_PER_BUCKET
    t = (
        "\\begin{table}[h!] \n \\centering \n \\setlength{\\tabcolsep}{3pt} \n \\footnotesize \n \\arrayrulecolor{black} \\\\% Bordures en bleu \n \\color{black} \n \\begin{tabular}{|c|ccc|ccc|ccc|} \n \\hline \n  &  $MILP_{1h}$    &  $LBBD_{1h}$   &  Heuristic \\\\\n Instance &  \\#Opt & \\#Feasible & $\\overline{Gap}$ & \\#Opt & \\#Feasible & $\\overline{Gap}$ &  \\#Feasible & $\\overline{Gap_{UB}}$ &  $\\overline{Gap_{LB}}$  \\\\ \\hline \n"
    )
    t += "\\begin{tabular}{|c|c c| c c| c c|} \n \\hline \n"
    t += " & \\multicolumn{2}{c|}{10 customers} & \\multicolumn{2}{c|}{20 customers} & \\multicolumn{2}{c|}{30 customers} \\\\\n"
    t += " & 1h & 4h  & 1h & 4h & 1h & 4h \\\\\n\\hline \n"
    t += (
        f"Average number of iterations & {aver_ite10_Short/d:.3f} & {aver_ite10_Long/d:.3f} "
        f"& {aver_ite20_Short/d:.3f} & {aver_ite20_Long/d:.3f} & {aver_ite30_Short/d:.3f} "
        f"& {aver_ite30_Long/d:.3f} \\\\\n"
        f"Average master solving time (s) & {aver_MSolving10_Short/d:.3f} & {aver_MSolving10_Long/d:.3f} "
        f"& {aver_MSolving20_Short/d:.3f} & {aver_MSolving20_Long/d:.3f} "
        f"& {aver_MSolving30_Short/d:.3f} & {aver_MSolving30_Long/d:.3f} \\\\\n"
        f"Average subproblem solving time (s) & {aver_Subsolving10_Short/d:.3f} "
        f"& {aver_Subsolving10_Long/d:.3f} & {aver_Subsolving20_Short/d:.3f} "
        f"& {aver_Subsolving20_Long/d:.3f} & {aver_Subsolving30_Short/d:.3f} "
        f"& {aver_Subsolving30_Long/d:.3f} \\\\\n"
        f"Average final number of benders cuts & {aver_BendersCuts10_Short/d:.3f} "
        f"& {aver_BendersCuts10_Long/d:.3f} & {aver_BendersCuts20_Short/d:.3f} "
        f"& {aver_BendersCuts20_Long/d:.3f} & {aver_BendersCuts30_Short/d:.3f} "
        f"& {aver_BendersCuts30_Long/d:.3f} \\\\\n"
        "\\hline \n\\end{tabular} \n\\end{table} \n"
    )
    return t


def _pick_ff_long_csv(csv_file: str) -> bool:
    """First pass: associate a long TL All/FF CSV with the folder (substring match)."""
    return csv_file == SOL_ALL_14400 or FF_SUFFIX_14400 in csv_file


def _row_is_long_ff_solution(csv_file: str) -> bool:
    """Row processing: treat as long-horizon All/FF only on exact suffix (legacy behavior)."""
    return csv_file == SOL_ALL_14400 or csv_file.endswith(FF_SUFFIX_14400)


def _pick_ff_short_csv(csv_file: str) -> bool:
    """First pass: short-horizon All/FF (substring tag in filename)."""
    return csv_file == SOL_ALL_3600 or SHORT_HEURISTIC_RUN_TAG in csv_file


def _row_is_short_ff_solution(csv_file: str) -> bool:
    """Row processing: short-horizon run only if exact legacy filename suffix."""
    return csv_file == SOL_ALL_3600 or csv_file.endswith(SHORT_HEURISTIC_RUN_FILE)


def _pick_cplex_long_csv(csv_file: str) -> bool:
    return csv_file == SOL_CPLEX_14400 or CPL_MARKER_LONG_RUN in csv_file


def _pick_cplex_short_csv(csv_file: str) -> bool:
    return csv_file == SOL_CPLEX_3600 or CPL_MARKER_IN_SHORT_NAME in csv_file


def _mean_relative_ub_gap_pct(reference_ub: list, variant_ub: list) -> float:
    """Average ((variant - reference) / reference) * 100 over paired rows."""
    n = len(variant_ub)
    if n == 0:
        return 0.0
    total = 0.0
    for i in range(n):
        total += (
            (variant_ub[i] - reference_ub[i]) / reference_ub[i]
        ) * 100
    return total / n


def _compare_ff_cpl_lowers(folder_path: str, ff_csv: str, cpl_csv: str, folder_label_for_log: str):
    """
    Align FF and CPLEX CSVs on the first column (instance id) and compare Lower bounds.
    Returns (nb_instances_ff_lower_ge_cpl, mean_lb_gap_percent) or (None, None) on failure.
    """
    try:
        df_ff = pd.read_csv(os.path.join(folder_path, ff_csv), delimiter=";")
        df_cpl = pd.read_csv(os.path.join(folder_path, cpl_csv), delimiter=";")
        inst_ff = df_ff.set_index(df_ff.columns[0])
        inst_cpl = df_cpl.set_index(df_cpl.columns[0])
        common_idx = inst_ff.index.intersection(inst_cpl.index)
        lower_ff = pd.to_numeric(inst_ff.loc[common_idx, "Lower"], errors="coerce")
        lower_cpl = pd.to_numeric(inst_cpl.loc[common_idx, "Lower"], errors="coerce")
        nb_ff_ge_cpl = (lower_cpl.isna() | (lower_ff >= lower_cpl)).sum()
        lb_gap_pct_series = 100 - lower_cpl * 100 / lower_ff
        lb_gap_pct_series = lb_gap_pct_series[
            lower_cpl.notna() & (lower_cpl != 0) & lower_ff.notna()
        ]
        gap_mean = (
            lb_gap_pct_series.mean() if not lb_gap_pct_series.empty else None
        )
        return nb_ff_ge_cpl, gap_mean
    except Exception as e:
        print(f"Erreur comparaison Lower FF/CPL dans {folder_label_for_log} : {e}")
        return None, None


def _accum_variant_row(
    csv_file: str,
    target_csv: str,
    bucket_idx: int,
    count_opt: int,
    mean_gap,
    upper_column,
    aver_opt,
    aver_gap,
    temp_ub_columns,
):
    """When csv_file matches a variant solver config, aggregate opt count, gap sum, UB column."""
    if csv_file != target_csv:
        return
    aver_opt[bucket_idx] += count_opt
    aver_gap[bucket_idx] += mean_gap
    temp_ub_columns[bucket_idx] = upper_column


def process_csv_files(base_directory, output_file="resultats_combinesV2.xlsx"):
    """Agrège les CSV de solutions par dossier d'instance vers un fichier Excel.

    Étapes principales pour chaque instance (parcours dans un ordre fixe pour le LaTeX) :
      lecture de tous les ``*.csv``, écriture d'une feuille Excel avec lignes récap ;
      enrichissement UB / déviation / gap UB–Lower pour les runs All/FF courts et longs ;
      cumuls pour CPLEX vs LBBD et impression des morceaux de tableaux LaTeX (tables 3–7).
    """
    meanShortCPLDiff=0
    meanLongCPLDiff=0
    wb = Workbook()
    wb.remove(wb.active)

    # Variables pour stats globales
    total_opt_plus = 0
    total_gap_diffs = []
    total_opt_plus_cpl = 0
    total_gap_diffs_cpl = []
    total_Feas_plus = 0
    processed_folders = 0
    aver_ite_long=0
    aver_ite_short=0
    aver_ite10_Short=0
    aver_ite10_Long=0
    aver_ite20_Short=0
    aver_ite20_Long=0
    aver_ite30_Short=0
    aver_ite30_Long=0
    aver_MSolving10_Short=0
    aver_MSolving10_Long=0
    aver_MSolving20_Short=0
    aver_MSolving20_Long=0
    aver_MSolving30_Short=0
    aver_MSolving30_Long=0
    aver_Subsolving10_Short=0
    aver_Subsolving10_Long=0
    aver_Subsolving20_Short=0
    aver_Subsolving20_Long=0
    aver_Subsolving30_Short=0
    aver_Subsolving30_Long=0
    aver_BendersCuts10_Short=0
    aver_BendersCuts10_Long=0
    aver_BendersCuts20_Short=0
    aver_BendersCuts20_Long=0
    aver_BendersCuts30_Short=0
    aver_BendersCuts30_Long=0

    aver_opt_all=[0,0,0]
    aver_gap_all=[0,0,0]
    aver_ubgap_all=[0,0,0]
    temp_ub_column_all=[[], [], []]
    aver_opt_ic=[0,0,0]
    aver_gap_ic=[0,0,0]
    aver_ubgpa_ic=[0,0,0]
    temp_ub_column_ic=[[], [], []]
    aver_opt_ms=[0,0,0]
    aver_gap_ms=[0,0,0]
    aver_ubgap_ms=[0,0,0]
    temp_ub_column_ms=[[], [], []]
    aver_opt_rs=[0,0,0]
    aver_gap_rs=[0,0,0]
    aver_ubgap_rs=[0,0,0]
    temp_ub_column_rs=[[], [], []]
    aver_opt_gap=[0,0,0]
    aver_gap_gap=[0,0,0]
    aver_ubgap_gap=[0,0,0]
    temp_ub_column_gap=[[], [], []]
    aver_opt_ff=[0,0,0]
    aver_gap_ff=[0,0,0]
    aver_ubgap_ff=[0,0,0]
    temp_ub_column_ff=[[], [], []]
    aver_gap_def=[0,0,0]
    aver_opt_def=[0,0,0]
    temp_ub_column_def=[[], [], []]
    aver_gap_def2=[0,0,0]
    aver_opt_def2=[0,0,0]
    temp_ub_column_def2=[[], [], []]
    
    avg_GAP_CPL=[]
    avg_GAP_FF=[]
    avg_GAP_FF_Short=[]
    avg_GAP_CPL_Short=[]
    latex_table_6 = "\\begin{table}[h!] \n \\small \n \\centering \n \\arrayrulecolor{black}  % Bordures en bleu \n \\color{black} \n \\begin{tabular}{|c|ccc||ccc|} \n \\hline \n  &  $MILP_{1h}$    &  $LBBD_{1h}$   &  \\# instances where LBBD  & $MILP_{4h}$    & $LBBD_{4h}$   &  \\# instances where LBBD \\\\ \n Instance &  $\\overline{\text{LB}}$ &    $\\overline{\text{LB}}$ &     finds a better LB & $\\overline{\text{LB}}$ &    $\\overline{\text{LB}}$ &  finds a better LB \\\\ \\hline \n"
    latex_table_5 = "\\begin{table}[h!] \n \\small \n \\centering \n \\begin{tabular}{|c|ccc|ccc|ccc|} \n \\hline \n  &  $MILP_{4h}$    &  $LBBD_{4h}$   &  Heuristic \\\\\n Instance &  \\#Opt & \\#Feasible & $\\overline{Gap}$ & \\#Opt & \\#Feasible & $\\overline{Gap}$ &  \\#Feasible & $\\overline{Gap_{UB}}$ &  $\\overline{Gap_{LB}}$  \\\\ \\hline \n"
    latex_table_4 = "\\begin{table}[h!] \n \\small \n \\centering \n \\begin{tabular}{|c|ccc|ccc|ccc|} \n \\hline \n  &  $MILP_{1h}$    &  $LBBD_{1h}$   &  Heuristic \\\\\n Instance &  \\#Opt & \\#Feasible & $\\overline{Gap}$ & \\#Opt & \\#Feasible & $\\overline{Gap}$ &  \\#Feasible & $\\overline{Gap_{UB}}$ &  $\\overline{Gap_{LB}}$  \\\\ \\hline \n"
    latex_table_7 = "\\begin{table}[h!] \n \\centering \n \\setlength{\\tabcolsep}{3pt} \n \\small \n \\arrayrulecolor{black}  % Bordures en bleu \n \\color{black} \n \\begin{tabular}{|c|ccc | ccc | ccc |} \n \\hline \n  &\\multicolumn{3}{c|}{10 customers} & \\multicolumn{3}{c|}{20 customers} & \\multicolumn{3}{c|}{30 customers} \\\\\n"
    latex_table_7+= "Method & \\#Opt & $\\overline{Gap}$ &  $\\overline{Gap_{UB}}$ &  \\#Opt  & $\\overline{Gap}$  & $\\overline{Gap_{UB}}$ &   \\#Opt & $\\overline{Gap}$ & $\\overline{Gap_{UB}}$  \\\\ \\hline \n"
    if set(INSTANCE_FOLDER_ORDER) != set(INSTANCE_FOLDER_META.keys()):
        raise ValueError(
            "INSTANCE_FOLDER_ORDER et INSTANCE_FOLDER_META doivent décrire exactement "
            "les mêmes dossiers."
        )
    variant_sheet_targets = (
        (SOL_IC_3600, aver_opt_ic, aver_gap_ic, temp_ub_column_ic),
        (SOL_MS_3600, aver_opt_ms, aver_gap_ms, temp_ub_column_ms),
        (SOL_RS_3600, aver_opt_rs, aver_gap_rs, temp_ub_column_rs),
        (SOL_GAP_VARIANT_3600, aver_opt_gap, aver_gap_gap, temp_ub_column_gap),
        (SOL_FF_VARIANT_3600, aver_opt_ff, aver_gap_ff, temp_ub_column_ff),
        (SOL_DEFAULT_BC1, aver_opt_def, aver_gap_def, temp_ub_column_def),
        (
            SOL_LEGACY_DEFAULT_BC1,
            aver_opt_def2,
            aver_gap_def2,
            temp_ub_column_def2,
        ),
    )

    # Parcourir tous les dossiers d'instances (ordre fixé pour les sorties LaTeX)
    for raw_folder_name in INSTANCE_FOLDER_ORDER:
        folder_path = os.path.join(base_directory, raw_folder_name)
        folder_label, customer_bucket_idx = INSTANCE_FOLDER_META[raw_folder_name]
        inst10 = customer_bucket_idx == 0
        inst20 = customer_bucket_idx == 1
        inst30 = customer_bucket_idx == 2
        n = customer_bucket_idx

        if not os.path.isdir(folder_path):
            continue

        csv_files = sorted(f for f in os.listdir(folder_path) if f.endswith(".csv"))
        roles = _resolve_solution_csv_filenames(csv_files)
        long_tl_file = roles.long_ff_csv
        short_tl_file = roles.short_ff_csv
        longCPL_file = roles.long_cplex_csv
        shortCPL_file = roles.short_cplex_csv

        folder_name = folder_label
        
        
        ws = wb.create_sheet(title=folder_name[:31])
        current_row = 1
        csv_stats = []

        cpl_opt = cpl_feas = 0
        cpl_gap_avg = 0.0
        ff_opt = ff_feas = 0
        ff_gap_avg = 0.0
        ub_small_count = 0
        ub_small_count_short = 0
        mean_deviation_for_ff = None
        mean_deviation_for_ff_short = None
        mean_gap_ub_lower_for_ff = None
        mean_gap_ub_lower_for_ff_short = None
        nb_lower_ff_sup_cpl_long = 0
        nb_lower_ff_sup_cpl_short = 0

        for csv_file in csv_files:
            csv_path = os.path.join(folder_path, csv_file)
            try:
                df = pd.read_csv(csv_path, delimiter=';')
            except Exception as e:
                print(f"Erreur en lisant {csv_path}: {e}")
                continue

            # Colonnes Upper + UB / déviation pour les fichiers All/FF (voir constantes SOL_*)
            is_long_ff_tl = _row_is_long_ff_solution(csv_file)
            is_short_ff_tl = _row_is_short_ff_solution(csv_file)
            upper_column, d_ub_long, d_ub_short = _inject_ub_derived_columns(
                folder_path, df, is_long_ff_tl, is_short_ff_tl
            )
            ub_small_count += d_ub_long
            ub_small_count_short += d_ub_short

            if current_row > 1:
                ws.append(["---"] * len(df.columns))
                current_row += 1

            ws.append([csv_file] + [""] * (len(df.columns) - 1))
            current_row += 1

            header = df.columns.tolist()
            ws.append(header)
            current_row += 1

            for _, row in df.iterrows():
                ws.append(row.tolist())
                current_row += 1

            # Statistiques (ajustées si nouvelles colonnes)
            count_opt = (df['Status'] == 'OPT').sum()
            if(is_long_ff_tl or is_short_ff_tl):
                count_feas = (df['Upper'] < 10000).sum()
            else:
                count_feas = count_opt + (df['Status'] == 'FEAS').sum()
                

            try:
                mean_iteration = df['Iteration'].mean()
                mean_time_masters = df['Total time'].mean()
                mean_gap = df[' Gap'].mean()*100 if ' Gap' in df.columns else df['Gap'].mean()
                mean_deviation = df['Deviation %'].mean() if 'Deviation %' in df.columns else None
                mean_gap_ub_lower = df['Gap UB-Lower %'].mean() if 'Gap UB-Lower %' in df.columns else None
            except Exception:
                mean_iteration = mean_time_masters = mean_gap = 0
                mean_deviation = None
                mean_gap_ub_lower = None

            stats_row = [
                f"Stat: {count_opt} OPT, {count_feas} FEAS",
                "",
                mean_iteration,
                "",
                "",
                "",
                mean_time_masters,
                "",
                "",
                "",
                "",
                f"Gap: {mean_gap:.6f}",
                f"Dev %: {mean_deviation:.2f}%" if mean_deviation is not None and not pd.isna(mean_deviation) else "",
                f"Gap UB-L: {mean_gap_ub_lower:.2f}%" if 'Gap UB-Lower %' in df.columns and mean_gap_ub_lower is not None else ""
            ]

            # Adapter la longueur selon les colonnes
            stats_row = stats_row[:len(df.columns)]
            ws.append(stats_row)
            current_row += 1

            csv_stats.append({
                'csv_name': csv_file,
                'count_opt': count_opt,
                'count_feas': count_feas,
                'mean_time_masters': mean_time_masters,
                'mean_gap': mean_gap,
                'mean_deviation': mean_deviation,
                'mean_gap_ub_lower': mean_gap_ub_lower,
                'is_long_tl': csv_file == long_tl_file,
                'is_short_tl': csv_file == short_tl_file,
                'is_long_cpl': csv_file == longCPL_file,
                'is_short_cpl': csv_file == shortCPL_file
            })
            for tgt, ao, ag, tu in variant_sheet_targets:
                _accum_variant_row(
                    csv_file, tgt, n, count_opt, mean_gap, upper_column,
                    ao, ag, tu,
                )
            # Remplir les stats pour la ligne LaTeX
            if csv_file == longCPL_file:
                cpl_opt = count_opt
                cpl_feas = count_feas
                cpl_gap_avg = mean_gap
                cpl_long_lb = df['Lower'].mean()
                avg_GAP_CPL.append(mean_gap)
            if csv_file == long_tl_file:
                ff_opt = count_opt
                ff_feas = count_feas
                ff_gap_avg = mean_gap
                avg_GAP_FF.append(mean_gap)
                mean_deviation_for_ff = mean_deviation
                mean_gap_ub_lower_for_ff = mean_gap_ub_lower
                aver_ite_long+= mean_iteration
                ff_long_lb = df['Lower'].mean()
                if inst10:
                    aver_ite10_Long+= mean_iteration
                    aver_BendersCuts10_Long+= (df['Nb Feas cut'].sum()+df[' Nb Opt cut '].sum())/10
                    aver_MSolving10_Long+= df['Time to solve Masters'].mean()
                    aver_Subsolving10_Long+= df['Time to solve Subs'].mean()
                if inst20:
                    aver_ite20_Long+= mean_iteration
                    aver_BendersCuts20_Long+= (df['Nb Feas cut'].sum()+df[' Nb Opt cut '].sum())/10
                    aver_MSolving20_Long+= df['Time to solve Masters'].mean()
                    aver_Subsolving20_Long+= df['Time to solve Subs'].mean()
                if inst30:
                    aver_ite30_Long+= mean_iteration
                    aver_BendersCuts30_Long+= (df['Nb Feas cut'].sum()+df[' Nb Opt cut '].sum())/10
                    aver_MSolving30_Long+= df['Time to solve Masters'].mean()
                    aver_Subsolving30_Long+= df['Time to solve Subs'].mean()
            if csv_file == shortCPL_file:
                cpl_short_lb = df['Lower'].mean()
                avg_GAP_CPL_Short.append(mean_gap)
                cpl_opt_short = count_opt
                cpl_feas_short = count_feas
                cpl_gap_avg_short = mean_gap
            if csv_file == short_tl_file:
                avg_GAP_FF_Short.append(mean_gap)
                ff_short_lb = df['Lower'].mean()
                aver_ite_short+= mean_iteration
                ff_opt_short = count_opt
                ff_feas_short = count_feas
                ff_gap_avg_short = mean_gap
                mean_deviation_for_ff_short = copy.deepcopy(mean_deviation)
                mean_gap_ub_lower_for_ff_short = copy.deepcopy(mean_gap_ub_lower)
                if inst10:
                    aver_ite10_Short+= mean_iteration
                    aver_BendersCuts10_Short+= (df['Nb Feas cut'].sum()+df[' Nb Opt cut '].sum())/10
                    aver_MSolving10_Short+= df['Time to solve Masters'].mean()
                    aver_Subsolving10_Short+= df['Time to solve Subs'].mean()

                    temp_ub_column_all[0]=upper_column
                    aver_gap_all[0]+=mean_gap
                    aver_opt_all[0]+=count_opt
                if inst20:
                    aver_ite20_Short+= mean_iteration
                    aver_BendersCuts20_Short+= (df['Nb Feas cut'].sum()+df[' Nb Opt cut '].sum())/10
                    aver_MSolving20_Short+= df['Time to solve Masters'].mean()
                    aver_Subsolving20_Short+= df['Time to solve Subs'].mean()
                    temp_ub_column_all[1]=upper_column
                    aver_gap_all[1]+=mean_gap
                    aver_opt_all[1]+=count_opt
                if inst30:
                    aver_ite30_Short+= mean_iteration
                    aver_BendersCuts30_Short+= (df['Nb Feas cut'].sum()+df[' Nb Opt cut '].sum())/10
                    aver_MSolving30_Short+= df['Time to solve Masters'].mean()
                    aver_Subsolving30_Short+= df['Time to solve Subs'].mean()
                    temp_ub_column_all[2]=upper_column
                    aver_gap_all[2]+=mean_gap
                    aver_opt_all[2]+=count_opt

       
        if long_tl_file and longCPL_file:
            nb_l, gap_m = _compare_ff_cpl_lowers(
                folder_path, long_tl_file, longCPL_file, folder_name
            )
            if nb_l is not None:
                nb_lower_ff_sup_cpl_long = nb_l
            if gap_m is not None:
                meanLongCPLDiff += gap_m
        if short_tl_file and shortCPL_file:
            nb_s, gap_ms = _compare_ff_cpl_lowers(
                folder_path, short_tl_file, shortCPL_file, folder_name
            )
            if nb_s is not None:
                nb_lower_ff_sup_cpl_short = nb_s
            if gap_ms is not None:
                meanShortCPLDiff += gap_ms

        ws.append([])
        current_row += 1
        ws.append(["Recap"])
        current_row += 1

        recap_header = ["CSV", "OPT", "FEAS", "Time Masters (moy)", "Gap (moy)", "Dev % (moy)", "Gap UB-Lower %"]
        ws.append(recap_header)
        current_row += 1

        for stat in csv_stats:
            recap_row = [
                stat['csv_name'][:20],
                stat['count_opt'],
                stat['count_feas'],
                round(stat['mean_time_masters'], 2),
                f"{stat['mean_gap']:.6f}",
                f"{stat['mean_deviation']:.2f}%" if stat['mean_deviation'] is not None else "",
                f"{stat['mean_gap_ub_lower']:.2f}%" if stat.get('mean_gap_ub_lower') is not None else ""
            ]
            ws.append(recap_row)
            current_row += 1

        if long_tl_file and short_tl_file:
            long_stats = next(s for s in csv_stats if s['is_long_tl'])
            short_stats = next(s for s in csv_stats if s['is_short_tl'])

            opt_plus = long_stats['count_opt'] - short_stats['count_opt']
            if short_stats['mean_gap'] != 0:
                gap_diff_pct = ((long_stats['mean_gap'] - short_stats['mean_gap']) / short_stats['mean_gap']) * 100
            else:
                gap_diff_pct = 0

            total_opt_plus += opt_plus
            total_gap_diffs.append(gap_diff_pct)
            processed_folders += 1

            ws.append([])
            current_row += 1
            ws.append(["COMPARAISON TL=14400 vs TL=14400"])
            current_row += 1
            ws.append(["OPT en plus (TL=14400)", f"{opt_plus} OPT"])
            current_row += 1
            ws.append(["Écart moyennes GAP (%)", f"{gap_diff_pct:.2f}%"])
            current_row += 1

        if longCPL_file and shortCPL_file:
            long_stats = next(s for s in csv_stats if s['is_long_cpl'])
            short_stats = next(s for s in csv_stats if s['is_short_cpl'])

            opt_plus = long_stats['count_opt'] - short_stats['count_opt']
            feas_plus = long_stats['count_feas'] - short_stats['count_feas']
            if short_stats['mean_gap'] != 0:
                gap_diff_pct = ((long_stats['mean_gap'] - short_stats['mean_gap']) / short_stats['mean_gap']) * 100
            else:
                gap_diff_pct = 0

            total_opt_plus_cpl += opt_plus
            total_Feas_plus += feas_plus
            total_gap_diffs_cpl.append(gap_diff_pct)

        _style_instance_sheet(ws)

        if long_tl_file:
            mean_deviation_for_ff= format_pct(mean_deviation_for_ff)
            mean_gap_ub_lower_for_ff = format_pct(mean_gap_ub_lower_for_ff)
            cpl_gap_avg  = format_pct(cpl_gap_avg)
            ff_gap_avg = format_pct(ff_gap_avg)
            dev_str = f"{mean_deviation_for_ff:.3f}" if mean_deviation_for_ff is not None else "-"
            gapub_str = f"{mean_gap_ub_lower_for_ff:.3f}" if mean_gap_ub_lower_for_ff is not None else "-"
            latex_table_5+=f"{folder_name} & {cpl_opt} & {cpl_feas} & {cpl_gap_avg:.3f}\\% & "
            latex_table_5+=f"{ff_opt} & {ff_feas} & {ff_gap_avg:.3f}\\% & "
            latex_table_5+=f"{ub_small_count} & {dev_str}\\% & {gapub_str}\\% \\\\ \n"
        

        if short_tl_file:
            mean_deviation_for_ff_short= format_pct(mean_deviation_for_ff_short)
            mean_gap_ub_lower_for_ff_short = format_pct(mean_gap_ub_lower_for_ff_short)
            cpl_gap_avg  = format_pct(cpl_gap_avg)
            ff_gap_avg = format_pct(ff_gap_avg)
            dev_str_short = f"{mean_deviation_for_ff_short:.3f}" if mean_deviation_for_ff_short is not None else "-"
            gapub_str = f"{mean_gap_ub_lower_for_ff_short:.3f}" if mean_gap_ub_lower_for_ff_short is not None else "-"
            latex_table_4+=f"{folder_name} & {cpl_opt_short} & {cpl_feas_short} & {cpl_gap_avg_short:.3f}\\ \\% & "
            latex_table_4+=f"{ff_opt_short} & {ff_feas_short} & {ff_gap_avg_short:.3f}\\ \\% & "
            latex_table_4+=f"{ub_small_count_short} & {dev_str_short}\\ \\% & {gapub_str}\\ \\% \\\\ \n"

            ref_ub = temp_ub_column_all[n]
            aver_ubgpa_ic[n] += _mean_relative_ub_gap_pct(ref_ub, temp_ub_column_ic[n])
            aver_ubgap_ms[n] += _mean_relative_ub_gap_pct(ref_ub, temp_ub_column_ms[n])
            aver_ubgap_rs[n] += _mean_relative_ub_gap_pct(ref_ub, temp_ub_column_rs[n])
            aver_ubgap_gap[n] += _mean_relative_ub_gap_pct(ref_ub, temp_ub_column_gap[n])
            aver_ubgap_ff[n] += _mean_relative_ub_gap_pct(ref_ub, temp_ub_column_ff[n])
            aver_ubgap_all[n] += _mean_relative_ub_gap_pct(ref_ub, ref_ub)
    
        if longCPL_file:
            mean_deviation_for_ff= format_pct(mean_deviation_for_ff)
            mean_gap_ub_lower_for_ff = format_pct(mean_gap_ub_lower_for_ff)
            cpl_gap_avg  = format_pct(cpl_gap_avg)
            ff_gap_avg = format_pct(ff_gap_avg)

        if long_tl_file :
            latex_table_6+=f"{folder_name} & {cpl_short_lb:.3f} & {ff_short_lb:.3f} & {nb_lower_ff_sup_cpl_short} & {cpl_long_lb:.3f} & {ff_long_lb:.3f} & {nb_lower_ff_sup_cpl_long} \\\\ \n"

    latex_table_7 = _append_latex_table7_aggregate_rows(
        latex_table_7,
        aver_opt_all=aver_opt_all,
        aver_gap_all=aver_gap_all,
        aver_ubgap_all=aver_ubgap_all,
        aver_opt_ic=aver_opt_ic,
        aver_gap_ic=aver_gap_ic,
        aver_ubgpa_ic=aver_ubgpa_ic,
        aver_opt_gap=aver_opt_gap,
        aver_gap_gap=aver_gap_gap,
        aver_ubgap_gap=aver_ubgap_gap,
        aver_opt_ms=aver_opt_ms,
        aver_gap_ms=aver_gap_ms,
        aver_ubgap_ms=aver_ubgap_ms,
        aver_opt_rs=aver_opt_rs,
        aver_gap_rs=aver_gap_rs,
        aver_ubgap_rs=aver_ubgap_rs,
        aver_opt_ff=aver_opt_ff,
        aver_gap_ff=aver_gap_ff,
        aver_ubgap_ff=aver_ubgap_ff,
        aver_opt_def=aver_opt_def,
        aver_gap_def=aver_gap_def,
    )
    latex_table_6 = _latex_close_outer_table(latex_table_6)
    latex_table_5 = _latex_close_outer_table(latex_table_5)
    latex_table_4 = _latex_close_outer_table(latex_table_4)

    wb.save(output_file)
    latex_table_3 = _build_latex_table3_runtime_totals(
        aver_ite10_Short=aver_ite10_Short,
        aver_ite10_Long=aver_ite10_Long,
        aver_ite20_Short=aver_ite20_Short,
        aver_ite20_Long=aver_ite20_Long,
        aver_ite30_Short=aver_ite30_Short,
        aver_ite30_Long=aver_ite30_Long,
        aver_MSolving10_Short=aver_MSolving10_Short,
        aver_MSolving10_Long=aver_MSolving10_Long,
        aver_MSolving20_Short=aver_MSolving20_Short,
        aver_MSolving20_Long=aver_MSolving20_Long,
        aver_MSolving30_Short=aver_MSolving30_Short,
        aver_MSolving30_Long=aver_MSolving30_Long,
        aver_Subsolving10_Short=aver_Subsolving10_Short,
        aver_Subsolving10_Long=aver_Subsolving10_Long,
        aver_Subsolving20_Short=aver_Subsolving20_Short,
        aver_Subsolving20_Long=aver_Subsolving20_Long,
        aver_Subsolving30_Short=aver_Subsolving30_Short,
        aver_Subsolving30_Long=aver_Subsolving30_Long,
        aver_BendersCuts10_Short=aver_BendersCuts10_Short,
        aver_BendersCuts10_Long=aver_BendersCuts10_Long,
        aver_BendersCuts20_Short=aver_BendersCuts20_Short,
        aver_BendersCuts20_Long=aver_BendersCuts20_Long,
        aver_BendersCuts30_Short=aver_BendersCuts30_Short,
        aver_BendersCuts30_Long=aver_BendersCuts30_Long,
    )
    
    print(latex_table_3)
    print(latex_table_4)
    print(latex_table_5)
    print(latex_table_6)
    print(latex_table_7)



    print(f"Combined result printed in: {output_file}")


if __name__ == "__main__":
    #Get current directory
    current_dir = os.getcwd()
    process_csv_files(current_dir, "combined_results.xlsx")
